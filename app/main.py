# Import FastAPI framework for building the web API
from fastapi import FastAPI, Request, Form, HTTPException
# Import StaticFiles for serving static files like CSS and JavaScript
from fastapi.staticfiles import StaticFiles
# Import Jinja2Templates for rendering HTML templates
from fastapi.templating import Jinja2Templates
# Import response classes for different types of HTTP responses
from fastapi.responses import RedirectResponse, StreamingResponse
# Import uvicorn for running the ASGI server
import uvicorn
# Import OS for file system operations
import os
# Import pandas for data manipulation and analysis
import pandas as pd
# Import openpyxl for Excel file creation and manipulation
from openpyxl import Workbook
# Import styling classes from openpyxl for Excel formatting
from openpyxl.styles import Font, PatternFill, Alignment
# Import utility function to convert column number to letter
from openpyxl.utils import get_column_letter
# Import BytesIO for creating in-memory binary streams
from io import BytesIO
# Import datetime for timestamp operations
from datetime import datetime

# ============================================================================
# SOLANA CONFIGURATION
# ============================================================================
SOLANA_ADDRESS = "FUmpfcaHxc6w8e2WJZrGMaWdBoDJF1NshTT7GesQAQft"

# Import routers (assuming they exist or will be created)
# try:
#     from auth.auth_routes import router as auth_router
# except ImportError:
#     auth_router = None

# try:
#     from logistics.logistics_routes import router as logistics_router
# except ImportError:
#     logistics_router = None

# try:
#     from realtime.realtime_routes import router as realtime_router
# except ImportError:
#     realtime_router = None

# try:
#     from reports.report_routes import router as reports_router
# except ImportError:
#     reports_router = None

# try:
#     from users.user_routes import router as users_router
# except ImportError:
#     users_router = None

# try:
#     from blockchain.solana_transactions import router as blockchain_router
# except ImportError:
#     blockchain_router = None

# Set router variables to None since they are not implemented yet
auth_router = None
logistics_router = None
realtime_router = None
reports_router = None
users_router = None
blockchain_router = None

# Create FastAPI application instance with title and description
app = FastAPI(title="B To B INTELLICA", description="Automated system for sales representatives to track deliveries and generate reports")

# add session middleware so we can remember wallet connections
from starlette.middleware.sessions import SessionMiddleware

# secret should be replaced with a secure env variable in production
app.add_middleware(SessionMiddleware, secret_key=os.getenv("SESSION_SECRET", "super-secret-key"))

# Mount static files (configuring a web server or framework to serve non-changing assets)
# Get the path to the static directory
static_dir = os.path.join(os.path.dirname(__file__), "static")
# Check if static directory exists
if not os.path.exists(static_dir):
    # Print warning if static directory is not found
    print(f"Warning: Static directory not found: {static_dir}")
else:
    # Print confirmation if static directory is found
    print(f"Static directory found: {static_dir}")
    # Mount the static files directory to serve CSS, JS, and other static assets
    app.mount("/static", StaticFiles(directory=static_dir), name="static")

# Templates
# Get the path to the templates directory
templates_dir = os.path.join(os.path.dirname(__file__), "templates")
# Check if templates directory exists
if not os.path.exists(templates_dir):
    # Print warning if templates directory is not found
    print(f"Warning: Templates directory not found: {templates_dir}")
    # Set templates to None if directory doesn't exist
    templates = None
else:
    # Print confirmation if templates directory is found
    print(f"Templates directory found: {templates_dir}")
    # Initialize Jinja2Templates with the templates directory
    templates = Jinja2Templates(directory=templates_dir)

# Load dataset (used superamarket_data.csv as the dataset)
# Construct path to the supermarket data CSV file
data_path = os.path.join(os.path.dirname(__file__), "data", "supermarket_data.csv")
# Try to load the dataset
try:
    # Read the CSV file into a pandas DataFrame
    df = pd.read_csv(data_path)
    # Print success message with number of rows loaded
    print(f"Dataset loaded successfully: {len(df)} rows")
except Exception as e:
    # Print error message if loading fails
    print(f"Error loading dataset: {e}")
    # Set df to None if loading fails
    df = None

# Include routers if they exist
# if auth_router:
#     app.include_router(auth_router, prefix="/auth", tags=["Authentication"])
# if logistics_router:
#     app.include_router(logistics_router, prefix="/logistics", tags=["Logistics"])
# if realtime_router:
#     app.include_router(realtime_router, prefix="/realtime", tags=["Real-time"])
# if reports_router:
#     app.include_router(reports_router, prefix="/reports", tags=["Reports"])
# if users_router:
#     app.include_router(users_router, prefix="/users", tags=["Users"])
# if blockchain_router:
#     app.include_router(blockchain_router, prefix="/blockchain", tags=["Blockchain"])

# Import TrackingService for blockchain tracking endpoints
from app.blockchain.tracking_service import TrackingService

# Homepage route
# Define GET endpoint for the homepage
@app.get("/")
# Define async function to handle homepage requests
async def homepage(request: Request):
    # if user not connected, redirect to login
    if not request.session.get("wallet"):
        return RedirectResponse(url="/login")
    # Check if templates are loaded
    if templates is None:
        # Return error response if templates are not available
        return {"error": "Templates not loaded"}
    # Define list of available services with names and URLs
    services = [
        {"name": "Real-time Delivery Tracking", "url": "/tracking"},
        {"name": "Generate Real-time Report", "url": "/reports/generate"},
        {"name": "View Previous Reports", "url": "/reports/view"},
        {"name": "Quantity Confirmation", "url": "/quantity"},
        {"name": "Quality Confirmation", "url": "/quality"},
        {"name": "Blockchain Tracking", "url": "/blockchain/tracking"}
    ]
    # Render and return the homepage template with services data
    return templates.TemplateResponse("homepage.html", {"request": request, "services": services})

# Login page route
@app.get("/login")
async def login_page(request: Request):
    if templates is None:
        return {"error": "Templates not loaded"}
    return templates.TemplateResponse("login.html", {"request": request})

# Wallet login endpoint
@app.post("/wallet_login")
async def wallet_login(request: Request):
    data = await request.json()
    pubkey = data.get("public_key")
    if not pubkey:
        raise HTTPException(status_code=400, detail="public_key required")
    request.session["wallet"] = pubkey
    return {"status": "ok"}

# Logout endpoint
@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login")

# Handle service selection
# Define POST endpoint for service selection
@app.post("/select_service")
# Define async function to handle service selection form submission
async def select_service(request: Request, service: str = Form(...)):
    # Create dictionary mapping service names to their URLs
    service_urls = {
        "Real-time Delivery Tracking": "/tracking",
        "Generate Real-time Report": "/reports/generate",
        "View Previous Reports": "/reports/view",
        "Quantity Confirmation": "/quantity",
        "Quality Confirmation": "/quality",
        "Blockchain Tracking": "/blockchain/tracking"
    }
    # Check if selected service exists in the URL mapping
    if service in service_urls:
        # Redirect to the appropriate service URL
        return RedirectResponse(url=service_urls[service], status_code=303)
    else:
        # Raise HTTP exception for invalid service selection
        raise HTTPException(status_code=400, detail="Invalid service selected")

# API endpoint for recent deliveries
# Define GET endpoint for retrieving recent deliveries data
@app.get("/api/deliveries/recent")
# Define async function to handle recent deliveries API requests
async def get_recent_deliveries(limit: int = 20, offset: int = 0):
    # Check if dataset is loaded
    if df is None:
        # Raise HTTP exception if dataset is not available
        raise HTTPException(status_code=500, detail="Dataset not loaded")
    
    # Check if this is the first page (offset = 0)
    if offset == 0:
        # Get the most recent 'limit' number of deliveries
        recent = df.tail(limit)
    else:
        # Calculate start and end indices for pagination
        start_idx = -(offset + limit)
        end_idx = -offset if offset > 0 else None
        # Get deliveries for the specified page
        recent = df.iloc[start_idx:end_idx]
    
    # Select specific columns from the recent deliveries
    deliveries = recent[['Date', 'Time', 'Invoice ID', 'Branch', 'Product line', 'Quantity', 'Sales']].to_dict('records')
    # Return deliveries data as JSON response
    return {"deliveries": deliveries}

# API endpoint for all quantities
# Define GET endpoint for retrieving all quantities data
@app.get("/api/quantities/all")
# Define async function to handle all quantities API requests
async def get_all_quantities():
    # Check if dataset is loaded
    if df is None:
        # Raise HTTP exception if dataset is not available
        raise HTTPException(status_code=500, detail="Dataset not loaded")
    
    # Select specific columns for quantities data
    quantities = df[['Invoice ID', 'Quantity', 'Product line', 'Branch', 'Sales']].to_dict('records')
    # Return quantities data as JSON response
    return {"quantities": quantities}

# Tracking page (placeholder)
# Define GET endpoint for the tracking page
@app.get("/tracking")
# Define async function to handle tracking page requests
async def tracking(request: Request):
    # Check if user is authenticated
    if not request.session.get("wallet"):
        return RedirectResponse(url="/login")
    # Check if templates are loaded
    if templates is None:
        return {"error": "Templates not loaded"}
    # Load initial 3 deliveries server-side
    # Initialize empty list for initial deliveries
    initial_deliveries = []
    # Check if dataset is loaded
    if df is not None:
        # Get the 3 most recent deliveries
        recent = df.tail(3)
        # Convert to dictionary format for template
        initial_deliveries = recent[['Date', 'Time', 'Invoice ID', 'Branch', 'Product line', 'Quantity', 'Sales']].to_dict('records')
    # Render and return the tracking template with initial deliveries
    return templates.TemplateResponse("tracking.html", {"request": request, "initial_deliveries": initial_deliveries})

# Quantity Confirmation page
# Define GET endpoint for quantity confirmation page
@app.get("/quantity")
# Define async function to handle quantity confirmation page requests
async def quantity_confirmation(request: Request):
    # Check if user is authenticated
    if not request.session.get("wallet"):
        return RedirectResponse(url="/login")
    # Check if templates are loaded
    if templates is None:
        return {"error": "Templates not loaded"}
    # Check if dataset is loaded
    if df is None:
        # Return error template if dataset is not available
        return templates.TemplateResponse("quantity.html", {"request": request, "error": "Dataset not loaded", "invoice_ids": [], "productlines": []})
    
    # Get unique invoice IDs from the dataset
    invoice_ids = sorted(df['Invoice ID'].unique().tolist())
    # Get unique product lines from the dataset
    product_lines_unique = df['Product line'].unique()
    productlines = sorted([str(line) for line in product_lines_unique])
    # Render and return quantity confirmation template with invoice IDs and product lines
    return templates.TemplateResponse("quantity.html", {"request": request, "invoice_ids": invoice_ids, "productlines": productlines})

# Quality Confirmation page
# Define GET endpoint for quality confirmation page
@app.get("/quality")
# Define async function to handle quality confirmation page requests
async def quality_confirmation(request: Request):
    # Check if user is authenticated
    if not request.session.get("wallet"):
        return RedirectResponse(url="/login")
    # Check if templates are loaded
    if templates is None:
        return {"error": "Templates not loaded"}
    # Check if dataset is loaded
    if df is None:
        # Return error template if dataset is not available
        return templates.TemplateResponse("quality.html", {"request": request, "error": "Dataset not loaded"})
    
    # Group by Product line and calculate average rating as quality
    # Group data by product line and calculate mean rating
    quality_data = df.groupby('Product line')['Rating'].mean().reset_index()
    # Sort quality data by rating in descending order (highest rated first)
    quality_data = quality_data.sort_values('Rating', ascending=False)
    # Convert to list of dictionaries for template
    quality_list = quality_data.to_dict('records')
    # Render and return quality confirmation template with quality data
    return templates.TemplateResponse("quality.html", {"request": request, "quality_data": quality_list})

# Reports generate - Excel file
# Define GET endpoint for generating Excel reports
@app.get("/reports/generate")
# Define async function to handle report generation requests
async def generate_report(request: Request):
    # Check if user is authenticated
    if not request.session.get("wallet"):
        raise HTTPException(status_code=401, detail="Unauthorized - please login first")
    # Use try-except block to handle potential errors
    try:
        # Check if dataset is loaded
        if df is None:
            # Raise HTTP exception if dataset is not available
            raise HTTPException(status_code=500, detail="Dataset not loaded")
        
        # Create a new workbook
        wb = Workbook()
        
        # Summary Sheet
        # Get the active worksheet (first sheet)
        ws_summary = wb.active
        # Set the title of the summary sheet
        ws_summary.title = "Summary Report"
        
        # Styling
        # Define header font style (bold, white text)
        header_font = Font(bold=True, color="FFFFFF")
        # Define header fill style (indigo background)
        header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
        # Define center alignment style
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Summary headers
        # Set main title in cell A1
        ws_summary['A1'] = "B To B INTELLICA - Real-time Report"
        # Apply styling to the main title
        ws_summary['A1'].font = Font(bold=True, size=16, color="6366f1")
        # Merge cells A1 to E1 for the title
        ws_summary.merge_cells('A1:E1')
        
        # Set report generation timestamp
        ws_summary['A3'] = "Report Generated:"
        ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Summary data
        # Create list of summary metrics and values
        summary_data = [
            ["Metric", "Value"],
            ["Total Sales", f"${df['Sales'].sum():.2f}"],
            ["Total Quantity Delivered", int(df['Quantity'].sum())],
            ["Average Customer Rating", f"{df['Rating'].mean():.2f}"],
            ["Total Transactions", len(df)],
            ["Date Range", f"{df['Date'].min()} to {df['Date'].max()}"]
        ]
        
        # Write summary data to worksheet starting from row 5
        for row_num, row_data in enumerate(summary_data, 5):
            for col_num, cell_value in enumerate(row_data, 1):
                # Create cell with value
                cell = ws_summary.cell(row=row_num, column=col_num, value=cell_value)
                # Apply header styling to the first row of data
                if row_num == 5:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                # Apply center alignment to all cells
                cell.alignment = center_align
        
        # Sales by Product Line Sheet
        # Create new worksheet for product sales analysis
        ws_products = wb.create_sheet("Sales by Product")
        # Set title for the product sales sheet
        ws_products['A1'] = "Sales by Product Line"
        # Apply styling to the title
        ws_products['A1'].font = Font(bold=True, size=14, color="6366f1")
        # Merge cells for the title
        ws_products.merge_cells('A1:C1')
        
        # Group data by product line and sum sales
        product_sales = df.groupby('Product line')['Sales'].sum().reset_index()
        # Sort by sales in descending order
        product_sales = product_sales.sort_values('Sales', ascending=False)
        
        # Headers
        # Define column headers for product sales
        headers = ["Product Line", "Total Sales", "Percentage of Total"]
        # Write headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws_products.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        # Data
        # Calculate total sales for percentage calculation
        total_sales = df['Sales'].sum()
        # Write product sales data
        for row_num, (_, row) in enumerate(product_sales.iterrows(), 4):
            # Write product line name
            ws_products.cell(row=row_num, column=1, value=row['Product line'])
            # Write total sales with currency formatting
            ws_products.cell(row=row_num, column=2, value=f"${row['Sales']:.2f}")
            # Calculate and write percentage of total sales
            percentage = (row['Sales'] / total_sales) * 100
            ws_products.cell(row=row_num, column=3, value=f"{percentage:.1f}%")
        
        # Quantity by Branch Sheet
        # Create new worksheet for branch deliveries analysis
        ws_branches = wb.create_sheet("Deliveries by Branch")
        # Set title for the branch deliveries sheet
        ws_branches['A1'] = "Successful Deliveries by Branch"
        # Apply styling to the title
        ws_branches['A1'].font = Font(bold=True, size=14, color="6366f1")
        # Merge cells for the title
        ws_branches.merge_cells('A1:C1')
        
        # Group by branch and aggregate quantity and sales
        branch_deliveries = df.groupby('Branch').agg({
            'Quantity': 'sum',
            'Sales': 'sum',
            'Invoice ID': 'count'
        }).reset_index()
        # Sort by quantity delivered in descending order
        branch_deliveries = branch_deliveries.sort_values('Quantity', ascending=False)
        
        # Headers
        # Define column headers for branch deliveries
        headers = ["Branch", "Total Quantity Delivered", "Total Sales", "Number of Deliveries"]
        # Write headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws_branches.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        # Data
        # Write branch deliveries data
        for row_num, (_, row) in enumerate(branch_deliveries.iterrows(), 4):
            # Write branch name
            ws_branches.cell(row=row_num, column=1, value=row['Branch'])
            # Write total quantity delivered
            ws_branches.cell(row=row_num, column=2, value=int(row['Quantity']))
            # Write total sales with currency formatting
            ws_branches.cell(row=row_num, column=3, value=f"${row['Sales']:.2f}")
            # Write number of deliveries (count of invoice IDs)
            ws_branches.cell(row=row_num, column=4, value=int(row['Invoice ID']))
        
        # Recent Deliveries Sheet
        # Create new worksheet for recent deliveries
        ws_recent = wb.create_sheet("Recent Deliveries")
        # Set title for the recent deliveries sheet
        ws_recent['A1'] = "Recent Successful Deliveries"
        # Apply styling to the title
        ws_recent['A1'].font = Font(bold=True, size=14, color="6366f1")
        # Merge cells for the title
        ws_recent.merge_cells('A1:G1')
        
        # Get the 20 most recent deliveries
        recent_deliveries = df.tail(20)[['Date', 'Time', 'Invoice ID', 'Branch', 'Product line', 'Quantity', 'Sales']].copy()
        
        # Headers
        # Define column headers for recent deliveries
        headers = ["Date", "Time", "Invoice ID", "Branch", "Product Line", "Quantity", "Sales Amount"]
        # Write headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws_recent.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        # Data
        # Write recent deliveries data
        for row_num, (_, row) in enumerate(recent_deliveries.iterrows(), 4):
            # Write date
            ws_recent.cell(row=row_num, column=1, value=str(row['Date']))
            # Write time
            ws_recent.cell(row=row_num, column=2, value=str(row['Time']))
            # Write invoice ID
            ws_recent.cell(row=row_num, column=3, value=row['Invoice ID'])
            # Write branch
            ws_recent.cell(row=row_num, column=4, value=row['Branch'])
            # Write product line
            ws_recent.cell(row=row_num, column=5, value=row['Product line'])
            # Write quantity
            ws_recent.cell(row=row_num, column=6, value=int(row['Quantity']))
            # Write sales amount with currency formatting
            ws_recent.cell(row=row_num, column=7, value=f"${row['Sales']:.2f}")
        
        # Auto-adjust column widths
        # Iterate through all worksheets
        for ws in [ws_summary, ws_products, ws_branches, ws_recent]:
            # Iterate through each column in the worksheet
            for column in ws.columns:
                # Find the maximum length of content in the column
                max_length = 0
                # Use get_column_letter to safely convert column number to letter
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                # Set adjusted width (add 2 for padding)
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        # Create in-memory buffer for the Excel file
        buffer = BytesIO()
        # Save workbook to buffer
        wb.save(buffer)
        # Reset buffer position to beginning
        buffer.seek(0)
        
        # Generate filename with timestamp
        # Create timestamp string for unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Create filename with timestamp
        filename = f"B_To_B_INTELLICA_Report_{timestamp}.xlsx"
        
        # Return Excel file
        # Return the Excel file as a streaming response
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        # Handle any exceptions during report generation
        raise HTTPException(status_code=500, detail=f"Error generating report: {str(e)}")

@app.get("/blockchain/tracking")
async def blockchain_tracking(request: Request):
    """Blockchain-based shipment tracking dashboard"""
    # Check if user is authenticated
    if not request.session.get("wallet"):
        return RedirectResponse(url="/login")
    # Check if templates are loaded
    if templates is None:
        return {"error": "Templates not loaded"}
    # Render and return the blockchain tracking template
    return templates.TemplateResponse("blockchain_tracking.html", {"request": request})

# Create shipment on-chain
@app.post("/api/blockchain/shipments/create")
async def create_blockchain_shipment(request: Request):
    """Create a shipment on-chain (store critical data)"""
    if not request.session.get("wallet"):
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    try:
        data = await request.json()
        shipment_id = data.get("shipment_id")
        product_line = data.get("product_line")
        quantity = data.get("quantity")
        destination = data.get("destination")
        
        # Create GPS hash for initial location
        gps_data = {
            "latitude": data.get("latitude", 0),
            "longitude": data.get("longitude", 0),
            "timestamp": int(datetime.now().timestamp())
        }
        gps_hash = TrackingService.store_gps_data(shipment_id, gps_data)
        
        return {
            "status": "success",
            "shipment_id": shipment_id,
            "gps_hash": gps_hash,
            "message": "Shipment ready for blockchain deployment",
            "blockchain_instruction": {
                "function": "create_shipment",
                "params": {
                    "shipment_id": shipment_id,
                    "destination": destination,
                    "product_line": product_line,
                    "quantity": quantity,
                    "gps_data_hash": gps_hash
                }
            }
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Log tracking event (hybrid model)
@app.post("/api/blockchain/shipments/{shipment_id}/track")
async def log_tracking_event(shipment_id: str, request: Request):
    """
    Log shipment tracking event with hybrid storage:
    - Event hash stored on-chain
    - Detailed data stored off-chain
    """
    if not request.session.get("wallet"):
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    try:
        data = await request.json()
        event_type = data.get("event_type")  # "pickup", "in_transit", "delivery"
        location = data.get("location")
        
        # Optional: environmental data
        gps_data = data.get("gps")
        temperature = data.get("temperature")
        humidity = data.get("humidity")
        
        # Log event with hybrid storage
        event_record = TrackingService.log_tracking_event(
            shipment_id=shipment_id,
            event_type=event_type,
            location=location,
            gps_data=gps_data,
            temperature=temperature,
            humidity=humidity
        )
        
        # Determine storage strategy
        on_chain_data = event_record["on_chain_data"]
        off_chain_data = {
            "gps": gps_data,
            "environmental": {
                "temperature": temperature,
                "humidity": humidity
            }
        }
        
        # Generate Solana Explorer link for event hash (account or transaction)
        from app.blockchain.config import BlockchainConfig
        explorer_url = BlockchainConfig.get_explorer_url(event_record["event_hash"])
        return {
            "status": "success",
            "event_hash": event_record["event_hash"],
            "event_type": event_type,
            "location": location,
            "blockchain_instruction": {
                "function": "log_tracking_event",
                "params": on_chain_data
            },
            "off_chain_storage": {
                "location": "IPFS/Arweave",
                "data_hash": event_record["event_hash"],
                "details": off_chain_data
            },
            "solana_explorer_url": explorer_url
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Confirm delivery with proof
@app.post("/api/blockchain/shipments/{shipment_id}/confirm-delivery")
async def confirm_blockchain_delivery(shipment_id: str, request: Request):
    """
    Confirm delivery on-chain with proof
    Stores hash on-chain, actual proof off-chain
    """
    if not request.session.get("wallet"):
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    try:
        data = await request.json()
        recipient_name = data.get("recipient_name")
        signature = data.get("signature")  # Base64 encoded signature
        photo_hash = data.get("photo_hash")  # IPFS hash of delivery proof photo
        
        # Create delivery proof
        delivery_proof = TrackingService.create_delivery_proof(
            shipment_id=shipment_id,
            recipient_name=recipient_name,
            signature=signature,
            photo_hash=photo_hash
        )
        
        # Generate Solana Explorer link for delivery hash (account or transaction)
        from app.blockchain.config import BlockchainConfig
        explorer_url = BlockchainConfig.get_explorer_url(delivery_proof["delivery_hash"])
        return {
            "status": "success",
            "delivery_hash": delivery_proof["delivery_hash"],
            "blockchain_instruction": {
                "function": "confirm_delivery",
                "params": delivery_proof["on_chain_data"]
            },
            "off_chain_proof": {
                "recipient_name": recipient_name,
                "signature_hash": signature,
                "photo_hash": photo_hash
            },
            "solana_explorer_url": explorer_url
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Get tracking history
@app.get("/api/blockchain/shipments/{shipment_id}/history")
async def get_tracking_history(shipment_id: str, request: Request):
    """Get complete on-chain tracking history for a shipment"""
    if not request.session.get("wallet"):
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    try:
        history = TrackingService.get_shipment_tracking_history(shipment_id)
        return {
            "status": "success",
            "shipment_id": shipment_id,
            "events_count": len(history),
            "events": history,
            "storage_model": {
                "on_chain": "Event hashes, status changes, delivery confirmations",
                "off_chain": "GPS sequences, sensor data, delivery proofs"
            }
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Verify data integrity
@app.post("/api/blockchain/verify/gps")
async def verify_gps_integrity(request: Request):
    """Verify GPS data integrity against stored hash"""
    if not request.session.get("wallet"):
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    try:
        data = await request.json()
        shipment_id = data.get("shipment_id")
        latitude = data.get("latitude")
        longitude = data.get("longitude")
        timestamp = data.get("timestamp")
        
        is_valid = TrackingService.verify_gps_data(
            shipment_id=shipment_id,
            latitude=latitude,
            longitude=longitude,
            timestamp=timestamp
        )
        
        return {
            "status": "success",
            "shipment_id": shipment_id,
            "integrity_verified": is_valid,
            "message": "GPS data integrity verified" if is_valid else "GPS data mismatch - integrity violation detected"
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Deploy shipment to blockchain (requires Phantom signature)
@app.post("/api/blockchain/deploy")
async def deploy_to_blockchain(request: Request):
    """
    Deploy prepared shipment transaction to Solana blockchain
    Requires Phantom wallet signature
    """
    if not request.session.get("wallet"):
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    try:
        data = await request.json()
        transaction_serialized = data.get("transaction")
        
        # In production: Sign and send to Solana network
        # For now: Return instructions for frontend to sign
        
        return {
            "status": "pending_signature",
            "message": "Transaction ready for Phantom wallet signature",
            "wallet": request.session.get("wallet"),
            "instructions": {
                "1": "Review transaction in Phantom wallet",
                "2": "Sign with your wallet",
                "3": "Transaction will be submitted to Solana Devnet"
            },
            "transaction": transaction_serialized
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Blockchain Result Pages
# Create Shipment Result Page
@app.get("/blockchain/shipment-result")
async def shipment_result(request: Request):
    """Display shipment creation result"""
    if not request.session.get("wallet"):
        return RedirectResponse(url="/login")
    if templates is None:
        return {"error": "Templates not loaded"}
    
    # Get shipment data from session
    shipment_data = request.session.get("shipment_result_data", {})
    return templates.TemplateResponse("shipment_result.html", {
        "request": request,
        "shipment_data": shipment_data
    })

# Log Tracking Event Result Page
@app.get("/blockchain/tracking-event-result")
async def tracking_event_result(request: Request):
    """Display tracking event logging result"""
    if not request.session.get("wallet"):
        return RedirectResponse(url="/login")
    if templates is None:
        return {"error": "Templates not loaded"}
    
    # Get event data from session
    event_data = request.session.get("event_result_data", {})
    return templates.TemplateResponse("tracking_event_result.html", {
        "request": request,
        "event_data": event_data
    })

# Confirm Delivery Result Page
@app.get("/blockchain/delivery-result")
async def delivery_result(request: Request):
    """Display delivery confirmation result"""
    if not request.session.get("wallet"):
        return RedirectResponse(url="/login")
    if templates is None:
        return {"error": "Templates not loaded"}
    
    # Get delivery data from session
    delivery_data = request.session.get("delivery_result_data", {})
    return templates.TemplateResponse("delivery_result.html", {
        "request": request,
        "delivery_data": delivery_data
    })

# Tracking History Result Page
@app.get("/blockchain/tracking-history-result")
async def tracking_history_result(request: Request):
    """Display tracking history result"""
    if not request.session.get("wallet"):
        return RedirectResponse(url="/login")
    if templates is None:
        return {"error": "Templates not loaded"}
    
    # Get history data from session
    history_data = request.session.get("history_result_data", {})
    return templates.TemplateResponse("tracking_history_result.html", {
        "request": request,
        "history_data": history_data
    })

# Verify GPS Result Page
@app.get("/blockchain/verify-gps-result")
async def verify_gps_result(request: Request):
    """Display GPS verification result"""
    if not request.session.get("wallet"):
        return RedirectResponse(url="/login")
    if templates is None:
        return {"error": "Templates not loaded"}
    
    # Get GPS verification data from session
    gps_data = request.session.get("gps_result_data", {})
    return templates.TemplateResponse("verify_gps_result.html", {
        "request": request,
        "gps_data": gps_data
    })

# API endpoints to store results and redirect
# Store shipment result
@app.post("/api/blockchain/store-shipment-result")
async def store_shipment_result(request: Request):
    """Store shipment result in session and return redirect URL"""
    try:
        data = await request.json()
        request.session["shipment_result_data"] = data
        return {"redirect_url": "/blockchain/shipment-result"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Store tracking event result
@app.post("/api/blockchain/store-event-result")
async def store_event_result(request: Request):
    """Store event result in session and return redirect URL"""
    try:
        data = await request.json()
        request.session["event_result_data"] = data
        return {"redirect_url": "/blockchain/tracking-event-result"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Store delivery result
@app.post("/api/blockchain/store-delivery-result")
async def store_delivery_result(request: Request):
    """Store delivery result in session and return redirect URL"""
    try:
        data = await request.json()
        request.session["delivery_result_data"] = data
        return {"redirect_url": "/blockchain/delivery-result"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Store history result
@app.post("/api/blockchain/store-history-result")
async def store_history_result(request: Request):
    """Store history result in session and return redirect URL"""
    try:
        data = await request.json()
        request.session["history_result_data"] = data
        return {"redirect_url": "/blockchain/tracking-history-result"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Store GPS verification result
@app.post("/api/blockchain/store-gps-result")
async def store_gps_result(request: Request):
    """Store GPS result in session and return redirect URL"""
    try:
        data = await request.json()
        request.session["gps_result_data"] = data
        return {"redirect_url": "/blockchain/verify-gps-result"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Reports view
# Define GET endpoint for viewing reports page
@app.get("/reports/view")
# Define async function to handle reports view requests
async def view_reports(request: Request):
    # Check if user is authenticated
    if not request.session.get("wallet"):
        return RedirectResponse(url="/login")
    # Check if templates are loaded
    if templates is None:
        return {"error": "Templates not loaded"}
    # Render and return the reports template
    return templates.TemplateResponse("reports.html", {"request": request})

# Main execution block
# Check if this script is being run directly (not imported)
if __name__ == "__main__":
    # Run the FastAPI application using uvicorn server
    uvicorn.run(app, host="0.0.0.0", port=8000)


